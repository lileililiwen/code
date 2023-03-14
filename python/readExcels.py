import os
import openpyxl

# 遍历目录中的所有Excel文件
def read_excel_files(path):
    for filename in os.listdir(path):
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            filepath = os.path.join(path, filename)
            process_excel(filepath)

# 处理Excel文件
def process_excel(filepath):
    wb = openpyxl.load_workbook(filepath)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        for row in sheet.iter_rows(min_row=2):
            # 读取Excel文件中的内容
            name = row[0].value
            age = row[1].value
            gender = row[2].value

            # 将内容写入到新的Excel文件中
            write_excel(name, age, gender)

# 将内容写入到新的Excel文件中
def write_excel(name, age, gender):
    # 实现自己的写Excel代码逻辑
    pass

if __name__ == '__main__':
    path = '/path/to/your/folder'
    read_excel_files(path)