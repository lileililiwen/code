import openpyxl
import datetime
from wordpress_xmlrpc import Client, WordPressPost
from wordpress_xmlrpc.methods.posts import NewPost

# 连接 WordPress XML-RPC API
wp_url = "http://example.com/xmlrpc.php"
wp_username = "admin"
wp_password = "password"
client = Client(wp_url, wp_username, wp_password)

# 打开 Excel 文件并获取工作表
wb = openpyxl.load_workbook('data.xlsx')
sheet = wb['Sheet1']

# 遍历 Excel 行，插入 WordPress posts
for i in range(2, sheet.max_row + 1):
    title = sheet.cell(row=i, column=1).value
    content = sheet.cell(row=i, column=2).value
    # 获取 Excel 中的日期并将其转换为 WordPress post 的日期格式
    excel_date = sheet.cell(row=i, column=3).value
    post_date = datetime.datetime.strptime(excel_date, '%Y-%m-%d %H:%M:%S').strftime('%Y-%m-%d %H:%M:%S')
    post = WordPressPost()
    post.title = title
    post.content = content
    post.post_status = 'publish'
    post.date = post_date
    # 发布 WordPress post
    post_id = client.call(NewPost(post))
    print("Post %s inserted." % post_id)

# 关闭 Excel 文件
wb.close()